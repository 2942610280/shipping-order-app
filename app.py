import streamlit as st
import pandas as pd
import os
import re
import zipfile
import tempfile
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from openpyxl.drawing.image import Image as XLImage
from typing import Optional, Dict, List, Tuple, Any
from io import BytesIO
import warnings
warnings.filterwarnings('ignore')
# 尝试导入PIL
try:
    from PIL import Image as PILImage
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False
# ==================== 页面配置 ====================
st.set_page_config(
    page_title="出货单生成器",
    page_icon="📦",
    layout="wide"
)
# ==================== 样式 ====================
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1f2937;
        text-align: center;
        margin-bottom: 1rem;
    }
    .sub-header {
        font-size: 1.1rem;
        color: #6b7280;
        text-align: center;
        margin-bottom: 2rem;
    }
    .success-box {
        padding: 1rem;
        background-color: #d1fae5;
        border-radius: 0.5rem;
        border-left: 4px solid #10b981;
    }
    .warning-box {
        padding: 1rem;
        background-color: #fef3c7;
        border-radius: 0.5rem;
        border-left: 4px solid #f59e0b;
    }
    .stDownloadButton > button {
        width: 100%;
        background-color: #7c3aed;
        color: white;
    }
</style>
""", unsafe_allow_html=True)
# ==================== 出货单生成器类 ====================
class ShippingOrderGenerator:
    """出货单生成器 - Streamlit版本"""
    IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp'}
    CHINESE_NUMBERS = ["一", "二", "三", "四", "五", "六", "七", "八", "九", "十",
                       "十一", "十二", "十三", "十四", "十五", "十六", "十七", "十八", "十九", "二十"]
    COLUMN_WIDTHS = [12, 20, 10, 10, 15, 8, 10, 15, 12, 25, 20]
    ROW_HEIGHT = 60
    IMAGE_COL_WIDTH = 10
    def __init__(self, main_df, sku_id_df, supplier_sku_df, sku_name_df, 
                 barcode_files_dict=None, image_files_dict=None):
        """
        初始化生成器
        
        参数:
            main_df: 主数据表DataFrame
            sku_id_df: SKU对应货品ID表DataFrame
            supplier_sku_df: 供应商SKU表DataFrame
            sku_name_df: SKU名称表DataFrame
            barcode_files_dict: 条码文件字典 {文件名: 文件内容bytes}
            image_files_dict: 图片文件字典 {文件名: 文件内容bytes}
        """
        self.main_df = main_df
        self.main_df['原始顺序'] = range(len(self.main_df))
        self.sku_id_df = sku_id_df
        self.supplier_sku_df = supplier_sku_df
        self.sku_name_df = sku_name_df
        self.barcode_files_dict = barcode_files_dict or {}
        self.image_files_dict = image_files_dict or {}
        # 数据缓存
        self._product_id_index: Dict[str, pd.Series] = {}
        self._supplier_sku_cache: Dict[str, Tuple[str, bool]] = {}
        self._image_cache: Dict[str, Optional[bytes]] = {}
        # 列索引映射
        self.col_mapping: Dict[str, int] = {}
        # 初始化
        self._identify_columns()
        self._build_product_id_index()
        self._build_supplier_cache()
        self._build_image_cache()
    def _safe_str(self, value: Any) -> str:
        if pd.isna(value):
            return ""
        return str(value).strip()
    def _safe_int(self, value: Any, default: int = 0) -> int:
        if pd.isna(value):
            return default
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return default
    def _identify_columns(self):
        """识别SKU对应货品ID表中的关键列"""
        column_patterns = {
            '货品id': ['货品id', '货品Id', '货品ID'],
            '货品编码': ['货品编码', 'sku', 'SKU'],
            '单套个数': ['单套个数', '单套数量'],
            '商品详情': ['商品详情备注', '商品详情', '详情备注']
        }
        for col_idx, col_name in enumerate(self.sku_id_df.columns):
            col_str = str(col_name).strip()
            col_lower = col_str.lower()
            for key, patterns in column_patterns.items():
                if key not in self.col_mapping:
                    for pattern in patterns:
                        if pattern.lower() in col_lower or pattern in col_str:
                            self.col_mapping[key] = col_idx
                            break
        self.col_mapping.setdefault('货品id', 1)
        if '商品详情' not in self.col_mapping and len(self.sku_id_df.columns) > 2:
            self.col_mapping['商品详情'] = 2
    def _build_product_id_index(self):
        """建立货品ID索引"""
        product_id_col = self.col_mapping.get('货品id', 1)
        for idx, row in self.sku_id_df.iterrows():
            if len(row) > product_id_col and pd.notna(row.iloc[product_id_col]):
                product_id = self._safe_str(row.iloc[product_id_col])
                try:
                    normalized_id = str(int(float(product_id)))
                    self._product_id_index[normalized_id] = row
                except (ValueError, TypeError):
                    pass
                self._product_id_index[product_id] = row
    def _build_supplier_cache(self):
        """建立供应商SKU缓存"""
        current_supplier = "其他供应商"
        for _, row in self.supplier_sku_df.iterrows():
            for cell in row:
                if pd.notna(cell):
                    cell_str = str(cell).strip()
                    if re.search(r'[\u4e00-\u9fff]', cell_str) or '供应商' in cell_str or '厂' in cell_str:
                        current_supplier = cell_str
                    else:
                        self._supplier_sku_cache[cell_str] = (current_supplier, True)
    def _build_image_cache(self):
        """建立图片缓存"""
        for filename, content in self.image_files_dict.items():
            name_without_ext = os.path.splitext(filename)[0].lower()
            self._image_cache[name_without_ext] = content
    def _get_row_by_product_id(self, product_id: Any) -> Optional[pd.Series]:
        product_id_str = self._safe_str(product_id)
        if not product_id_str:
            return None
        if product_id_str in self._product_id_index:
            return self._product_id_index[product_id_str]
        try:
            normalized_id = str(int(float(product_id_str)))
            return self._product_id_index.get(normalized_id)
        except (ValueError, TypeError):
            return None
    def _extract_sku_prefix(self, sku: Any) -> str:
        sku_str = self._safe_str(sku)
        return sku_str.split('-')[0] if '-' in sku_str else sku_str
    def _get_multiplier_from_sku(self, sku: str) -> Optional[int]:
        match = re.search(r'-(\d+)[Xx]$', sku)
        return int(match.group(1)) if match else None
    def calculate_total_quantity(self, sku: Any, sets: Any, product_id: Any) -> int:
        sets_int = self._safe_int(sets)
        if sets_int <= 0:
            return 0
        product_id_str = self._safe_str(product_id)
        if product_id_str and '单套个数' in self.col_mapping:
            row = self._get_row_by_product_id(product_id)
            if row is not None:
                unit_qty = self._safe_int(row.iloc[self.col_mapping['单套个数']])
                if unit_qty > 0:
                    return sets_int * unit_qty
        sku_str = self._safe_str(sku)
        if not sku_str and product_id_str and '货品编码' in self.col_mapping:
            row = self._get_row_by_product_id(product_id)
            if row is not None:
                sku_str = self._safe_str(row.iloc[self.col_mapping['货品编码']])
        if sku_str:
            multiplier = self._get_multiplier_from_sku(sku_str)
            if multiplier:
                return sets_int * multiplier
        return sets_int
    def get_product_name(self, sku_prefix: str) -> str:
        if not sku_prefix:
            return ""
        for _, row in self.sku_name_df.iterrows():
            if len(row) >= 2 and pd.notna(row.iloc[0]):
                if sku_prefix in str(row.iloc[0]):
                    return self._safe_str(row.iloc[1]) or sku_prefix
        return sku_prefix
    def get_product_details(self, product_id: Any) -> str:
        if '商品详情' not in self.col_mapping:
            return ""
        row = self._get_row_by_product_id(product_id)
        if row is not None:
            return self._safe_str(row.iloc[self.col_mapping['商品详情']])
        return ""
    def get_supplier_group(self, sku_prefix: str) -> Tuple[str, bool]:
        if not sku_prefix:
            return "其他供应商", False
        if sku_prefix in self._supplier_sku_cache:
            return self._supplier_sku_cache[sku_prefix]
        for cached_sku, result in self._supplier_sku_cache.items():
            if sku_prefix in cached_sku or cached_sku in sku_prefix:
                return result
        return "其他供应商", False
    def find_image_data(self, sku_prefix: str) -> Optional[bytes]:
        """查找图片数据"""
        if not sku_prefix:
            return None
        sku_lower = sku_prefix.lower()
        # 精确匹配
        if sku_lower in self._image_cache:
            return self._image_cache[sku_lower]
        # 模糊匹配
        for name, data in self._image_cache.items():
            if sku_lower in name or name in sku_lower:
                return data
        return None
    def find_barcode_data(self, product_id: Any) -> Tuple[Optional[bytes], Optional[str]]:
        """查找条码数据"""
        product_id_str = self._safe_str(product_id)
        if not product_id_str:
            return None, None
        for filename, content in self.barcode_files_dict.items():
            if product_id_str in filename:
                return content, filename
        return None, None
    def _process_image_data(self, image_data: bytes) -> Optional[BytesIO]:
        """处理图片数据"""
        if not image_data:
            return None
        try:
            if not PIL_AVAILABLE:
                return BytesIO(image_data)
            img = PILImage.open(BytesIO(image_data))
            if img.mode in ('RGBA', 'LA', 'P'):
                if img.mode == 'P':
                    img = img.convert('RGBA')
                background = PILImage.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'RGBA':
                    background.paste(img, mask=img.split()[-1])
                    img = background
                else:
                    img = img.convert('RGB')
            elif img.mode != 'RGB':
                img = img.convert('RGB')
            buffer = BytesIO()
            img.save(buffer, format='JPEG', quality=95)
            buffer.seek(0)
            return buffer
        except Exception:
            return BytesIO(image_data)
    def _insert_image(self, ws, row: int, col: int, image_data: bytes) -> bool:
        """插入图片到Excel"""
        processed = self._process_image_data(image_data)
        if not processed:
            return False
        try:
            img = XLImage(processed)
            cell_width_px = self.IMAGE_COL_WIDTH * 7
            cell_height_px = self.ROW_HEIGHT * 1.33
            scale = min((cell_width_px * 0.85) / img.width, (cell_height_px * 0.85) / img.height)
            img.width = int(img.width * scale)
            img.height = int(img.height * scale)
            x_offset = (cell_width_px - img.width) / 2 + 1 + (0.1 / 2.54 * 96)
            y_offset = (cell_height_px - img.height) / 2 + 1 + (0.1 / 2.54 * 96)
            img.anchor = f"{get_column_letter(col)}{row}"
            img.left = int(x_offset * 9525)
            img.top = int(y_offset * 9525)
            ws.add_image(img)
            return True
        except Exception:
            return False
    def process_order_data(self, store_data: pd.DataFrame) -> Tuple[Dict[str, List], List]:
        """处理店铺数据"""
        supplier_orders: Dict[str, List] = {}
        abnormal_orders: List = []
        col_names = {
            'product_id': next((c for c in ['货品Id', '货品id', '货品ID'] if c in store_data.columns), None),
            'sku': next((c for c in ['货品编码', 'SKU', 'sku'] if c in store_data.columns), None),
            'sets': next((c for c in ['发货数量', '套数'] if c in store_data.columns), None),
            'address': '仓库地址' if '仓库地址' in store_data.columns else None,
            'warehouse': '仓库名称' if '仓库名称' in store_data.columns else None
        }
        for _, row in store_data.sort_values('原始顺序').iterrows():
            product_id = row.get(col_names['product_id'], '') if col_names['product_id'] else ''
            sku = row.get(col_names['sku'], '') if col_names['sku'] else ''
            sets = row.get(col_names['sets'], 0) if col_names['sets'] else 0
            sku_prefix = self._extract_sku_prefix(sku)
            supplier, found = self.get_supplier_group(sku_prefix)
            order_data = {
                'SKU': sku,
                '商品名称': self.get_product_name(sku_prefix),
                '商品图片数据': self.find_image_data(sku_prefix),
                'SKU前缀': sku_prefix,
                '商品详情': self.get_product_details(product_id),
                '套数': self._safe_int(sets),
                '总数量': self.calculate_total_quantity(sku, sets, product_id),
                '货品id': product_id,
                '仓库地址': row.get(col_names['address'], '') if col_names['address'] else '',
                '仓库名称': row.get(col_names['warehouse'], '') if col_names['warehouse'] else '',
                '原始顺序': row.get('原始顺序', 0),
                'barcode_data': None,
                'barcode_filename': None
            }
            # 查找条码
            barcode_data, barcode_name = self.find_barcode_data(product_id)
            if barcode_data:
                order_data['barcode_data'] = barcode_data
                order_data['barcode_filename'] = barcode_name
            if found:
                supplier_orders.setdefault(supplier, []).append(order_data)
            else:
                abnormal_orders.append(order_data)
        return supplier_orders, abnormal_orders
    def merge_orders(self, orders: List[Dict]) -> List[Dict]:
        """合并相同仓库和货品ID的订单"""
        merged: Dict[str, Dict] = {}
        for order in sorted(orders, key=lambda x: x['原始顺序']):
            key = f"{order['仓库名称']}_{order['货品id']}"
            if key not in merged:
                merged[key] = order.copy()
            else:
                merged[key]['套数'] += order['套数']
                merged[key]['总数量'] += order['总数量']
        return sorted(merged.values(), key=lambda x: x['原始顺序'])
    def group_by_warehouse(self, orders: List[Dict]) -> List[Dict]:
        """按仓库分组订单"""
        groups: Dict[str, Dict] = {}
        for order in orders:
            warehouse = order['仓库名称']
            if warehouse not in groups:
                groups[warehouse] = {
                    'warehouse_name': warehouse,
                    'warehouse_address': order['仓库地址'],
                    'orders': [],
                    'min_order': order['原始顺序']
                }
            groups[warehouse]['orders'].append(order)
            groups[warehouse]['min_order'] = min(groups[warehouse]['min_order'], order['原始顺序'])
        return sorted(groups.values(), key=lambda x: x['min_order'])
    def create_excel(self, supplier: str, orders: List[Dict], is_abnormal: bool = False) -> Tuple[BytesIO, List[Tuple[str, bytes]]]:
        """创建Excel出货单，返回Excel数据和条码文件列表"""
        warehouse_groups = self.group_by_warehouse(orders)
        barcode_files = []  # [(filename, data), ...]
        wb = Workbook()
        ws = wb.active
        ws.title = ("异常订单" if is_abnormal else supplier)[:31]
        # 标题
        ws.merge_cells('A1:K1')
        title = ws['A1']
        title.value = f"{'异常订单' if is_abnormal else supplier} 出货单 - {datetime.now().strftime('%Y-%m-%d')}"
        title.font = Font(bold=True, size=14)
        title.alignment = Alignment(horizontal='center', vertical='center')
        # 表头
        headers = ['单号', 'SKU', '商品名称', '商品图片', '商品详情', '套数', '总数量', '货品id', '条码文件',
                   '仓库地址', '仓库名称']
        header_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        # 填充数据
        current_row = 4
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        for wh_idx, wh_info in enumerate(warehouse_groups):
            start_row = current_row
            for i, order in enumerate(wh_info['orders']):
                ws.cell(row=current_row, column=2, value=order['SKU'])
                ws.cell(row=current_row, column=3, value=order['商品名称'])
                ws.cell(row=current_row, column=5, value=order['商品详情'])
                ws.cell(row=current_row, column=6, value=order['套数'])
                ws.cell(row=current_row, column=7, value=order['总数量'])
                id_cell = ws.cell(row=current_row, column=8, value=order['货品id'])
                if isinstance(order['货品id'], (int, float)):
                    id_cell.number_format = '0'
                # 条码文件名
                if order['barcode_data'] and order['barcode_filename']:
                    new_name = f"{order['套数']}--{order['barcode_filename']}"
                    ws.cell(row=current_row, column=9, value=new_name)
                    barcode_files.append((new_name, order['barcode_data']))
                else:
                    ws.cell(row=current_row, column=9, value="无条码")
                if i == 0:
                    ws.cell(row=current_row, column=10, value=wh_info['warehouse_address'])
                    ws.cell(row=current_row, column=11, value=wh_info['warehouse_name'])
                current_row += 1
            end_row = current_row - 1
            order_num = f"第{self.CHINESE_NUMBERS[wh_idx]}单" if wh_idx < len(
                self.CHINESE_NUMBERS) else f"第{wh_idx + 1}单"
            ws.cell(row=start_row, column=1, value=order_num)
            if end_row > start_row:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                ws.merge_cells(start_row=start_row, start_column=10, end_row=end_row, end_column=10)
                ws.merge_cells(start_row=start_row, start_column=11, end_row=end_row, end_column=11)
        # 插入图片
        img_row = 4
        for wh_info in warehouse_groups:
            for order in wh_info['orders']:
                if order['商品图片数据']:
                    self._insert_image(ws, img_row, 4, order['商品图片数据'])
                img_row += 1
        # 设置格式
        for i, width in enumerate(self.COLUMN_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        for row in range(4, current_row):
            ws.row_dimensions[row].height = self.ROW_HEIGHT
            alignments = [
                Alignment(horizontal='center', vertical='center'),
                Alignment(horizontal='left', vertical='center', wrap_text=True),
                Alignment(horizontal='center', vertical='center', wrap_text=True),
                Alignment(horizontal='center', vertical='center'),
                Alignment(horizontal='center', vertical='center', wrap_text=True),
                Alignment(horizontal='center', vertical='center'),
                Alignment(horizontal='center', vertical='center'),
                Alignment(horizontal='center', vertical='center'),
                Alignment(horizontal='center', vertical='center'),
                Alignment(horizontal='justify', vertical='center', wrap_text=True),
                Alignment(horizontal='justify', vertical='center', wrap_text=True),
            ]
            for col, align in enumerate(alignments, 1):
                ws.cell(row=row, column=col).alignment = align
        # 边框
        for row in ws.iter_rows(min_row=3, max_row=current_row - 1, min_col=1, max_col=11):
            for cell in row:
                cell.border = thin_border
        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output, barcode_files
    def generate_all_orders(self, progress_callback=None) -> BytesIO:
        """生成所有出货单，返回ZIP文件"""
        # 创建临时目录
        with tempfile.TemporaryDirectory() as temp_dir:
            output_folder = os.path.join(temp_dir, f'出货单_{datetime.now().strftime("%Y%m%d_%H%M%S")}')
            os.makedirs(output_folder)
            # 确定店铺列
            store_col = next((c for c in ['店铺名称', '店铺', '店铺名', '店名'] if c in self.main_df.columns), None)
            if not store_col:
                st.error(f"❌ 未找到店铺列。可用列: {list(self.main_df.columns)}")
                return None
            stores = list(self.main_df.groupby(store_col))
            total_stores = len(stores)
            for store_idx, (store_name, store_data) in enumerate(stores):
                if progress_callback:
                    progress_callback((store_idx + 1) / total_stores, f"处理店铺: {store_name}")
                safe_name = re.sub(r'[\\/*?:"<>|]', "_", str(store_name))
                store_folder = os.path.join(output_folder, f"店铺_{safe_name}")
                os.makedirs(store_folder, exist_ok=True)
                supplier_orders, abnormal_orders = self.process_order_data(store_data)
                # 处理正常订单
                for supplier, orders in supplier_orders.items():
                    if not orders:
                        continue
                    safe_supplier = re.sub(r'[\\/*?:"<>|]', "_", str(supplier))
                    supplier_folder = os.path.join(store_folder, f"供应商_{safe_supplier}")
                    os.makedirs(supplier_folder, exist_ok=True)
                    merged = self.merge_orders(orders)
                    excel_data, barcode_files = self.create_excel(supplier, merged)
                    # 保存Excel
                    excel_path = os.path.join(supplier_folder, f"{supplier}_出货单.xlsx")
                    with open(excel_path, 'wb') as f:
                        f.write(excel_data.getvalue())
                    # 保存条码文件
                    if barcode_files:
                        barcode_folder = os.path.join(supplier_folder, "条码")
                        os.makedirs(barcode_folder, exist_ok=True)
                        for filename, data in barcode_files:
                            barcode_path = os.path.join(barcode_folder, filename)
                            with open(barcode_path, 'wb') as f:
                                f.write(data)
                # 处理异常订单
                if abnormal_orders:
                    abnormal_folder = os.path.join(store_folder, "异常订单")
                    os.makedirs(abnormal_folder, exist_ok=True)
                    merged = self.merge_orders(abnormal_orders)
                    excel_data, barcode_files = self.create_excel("异常订单", merged, True)
                    excel_path = os.path.join(abnormal_folder, "异常订单_出货单.xlsx")
                    with open(excel_path, 'wb') as f:
                        f.write(excel_data.getvalue())
                    if barcode_files:
                        barcode_folder = os.path.join(abnormal_folder, "条码")
                        os.makedirs(barcode_folder, exist_ok=True)
                        for filename, data in barcode_files:
                            barcode_path = os.path.join(barcode_folder, filename)
                            with open(barcode_path, 'wb') as f:
                                f.write(data)
            # 打包成ZIP
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for root, dirs, files in os.walk(output_folder):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arc_name = os.path.relpath(file_path, temp_dir)
                        zip_file.write(file_path, arc_name)
            zip_buffer.seek(0)
            return zip_buffer
# ==================== 主界面 ====================
def main():
    st.markdown('<p class="main-header">📦 出货单生成器</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">上传所需文件，自动生成按供应商分组的出货单</p>', unsafe_allow_html=True)
    # 创建标签页
    tab1, tab2 = st.tabs(["📤 上传文件", "📖 使用说明"])
    with tab1:
        col1, col2 = st.columns(2)
        with col1:
            st.subheader("📄 必需的Excel文件")
            main_file = st.file_uploader(
                "1️⃣ 入库单列表页CO明细分页导出.xlsx",
                type=["xlsx", "xls"],
                help="主数据表，包含订单信息",
                key="main"
            )
            sku_id_file = st.file_uploader(
                "2️⃣ SKU对应货品id表.xlsx",
                type=["xlsx", "xls"],
                help="SKU与货品ID的对应关系",
                key="sku_id"
            )
            supplier_sku_file = st.file_uploader(
                "3️⃣ 同一供应商的不同SKU.xlsx",
                type=["xlsx", "xls"],
                help="供应商与SKU的分组关系",
                key="supplier"
            )
            sku_name_file = st.file_uploader(
                "4️⃣ SKU对应商品名称.xlsx",
                type=["xlsx", "xls"],
                help="SKU与商品名称的对应关系",
                key="sku_name"
            )
        with col2:
            st.subheader("📁 可选的附件文件")
            barcode_files = st.file_uploader(
                "📎 条码PDF文件（可多选）",
                type=["pdf"],
                accept_multiple_files=True,
                help="上传条码PDF文件，文件名需包含货品ID",
                key="barcodes"
            )
            image_files = st.file_uploader(
                "🖼️ 商品图片（可多选）",
                type=["jpg", "jpeg", "png", "gif", "bmp", "webp"],
                accept_multiple_files=True,
                help="上传商品图片，文件名需包含SKU前缀",
                key="images"
            )
        st.divider()
        # 检查文件状态
        all_required = all([main_file, sku_id_file, supplier_sku_file, sku_name_file])
        if all_required:
            st.success("✅ 所有必需文件已上传！")
            # 显示文件信息
            with st.expander("📊 查看上传文件信息", expanded=False):
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("主数据表", f"{main_file.name[:20]}...")
                with col2:
                    st.metric("条码文件", f"{len(barcode_files)} 个")
                with col3:
                    st.metric("图片文件", f"{len(image_files)} 个")
                with col4:
                    st.metric("PIL支持", "✅" if PIL_AVAILABLE else "❌")
            # 生成按钮
            if st.button("🚀 开始生成出货单", type="primary", use_container_width=True):
                try:
                    # 读取Excel文件
                    with st.spinner("📖 正在读取Excel文件..."):
                        main_df = pd.read_excel(main_file)
                        sku_id_df = pd.read_excel(sku_id_file)
                        supplier_sku_df = pd.read_excel(supplier_sku_file, header=None)
                        sku_name_df = pd.read_excel(sku_name_file)
                    st.info(f"📊 读取到 {len(main_df)} 条订单数据")
                    # 预览数据
                    with st.expander("👀 预览主数据表（前10行）"):
                        st.dataframe(main_df.head(10), use_container_width=True)
                    # 处理条码和图片文件
                    with st.spinner("📁 正在处理附件文件..."):
                        barcode_dict = {}
                        for f in barcode_files:
                            barcode_dict[f.name] = f.read()
                            f.seek(0)
                        image_dict = {}
                        for f in image_files:
                            image_dict[f.name] = f.read()
                            f.seek(0)
                    # 创建生成器
                    generator = ShippingOrderGenerator(
                        main_df=main_df,
                        sku_id_df=sku_id_df,
                        supplier_sku_df=supplier_sku_df,
                        sku_name_df=sku_name_df,
                        barcode_files_dict=barcode_dict,
                        image_files_dict=image_dict
                    )
                    # 生成出货单
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    def update_progress(progress, text):
                        progress_bar.progress(progress)
                        status_text.text(text)
                    with st.spinner("⚙️ 正在生成出货单..."):
                        zip_data = generator.generate_all_orders(progress_callback=update_progress)
                    progress_bar.progress(1.0)
                    status_text.text("✅ 生成完成！")
                    if zip_data:
                        st.success("🎉 出货单生成完成！")
                        # 下载按钮
                        st.download_button(
                            label="📥 下载出货单（ZIP压缩包）",
                            data=zip_data,
                            file_name=f"出货单_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                            mime="application/zip",
                            use_container_width=True
                        )
                        st.balloons()
                except Exception as e:
                    st.error(f"❌ 处理出错: {str(e)}")
                    with st.expander("🔍 查看详细错误信息"):
                        st.exception(e)
        else:
            st.warning("⚠️ 请上传所有必需的Excel文件")
            # 显示缺少的文件
            missing = []
            if not main_file:
                missing.append("入库单列表页CO明细分页导出.xlsx")
            if not sku_id_file:
                missing.append("SKU对应货品id表.xlsx")
            if not supplier_sku_file:
                missing.append("同一供应商的不同SKU.xlsx")
            if not sku_name_file:
                missing.append("SKU对应商品名称.xlsx")
            for f in missing:
                st.markdown(f"- ❌ **{f}**")
    with tab2:
        st.markdown("""
        ### 📖 使用说明
        
        #### 第一步：准备文件
        1. **入库单列表页CO明细分页导出.xlsx** - 主数据表，包含订单、SKU、数量等信息
        2. **SKU对应货品id表.xlsx** - SKU与货品ID的对应关系表
        3. **同一供应商的不同SKU.xlsx** - 供应商与SKU的分组关系
        4. **SKU对应商品名称.xlsx** - SKU与商品名称的对应表
        
        #### 第二步：上传附件（可选）
        - **条码PDF文件** - 文件名需包含货品ID，程序会自动匹配
        - **商品图片** - 文件名需包含SKU前缀，程序会自动匹配并插入Excel
        
        #### 第三步：生成出货单
        1. 上传所有必需文件后，点击"开始生成出货单"按钮
        2. 等待处理完成
        3. 点击"下载"按钮获取ZIP压缩包
        
        #### 输出内容
        - 按**店铺**分文件夹
        - 每个店铺下按**供应商**分文件夹
        - 每个供应商文件夹包含：
          - Excel出货单（含商品图片）
          - 条码文件夹（包含重命名后的条码PDF）
        - 异常订单（无法匹配供应商的）单独生成
        
        ---
        
        ### ❓ 常见问题
        
        **Q: 为什么有些图片没有显示？**
        > A: 请确保图片文件名包含对应的SKU前缀，程序通过文件名匹配图片。
        
        **Q: 条码文件如何匹配？**
        > A: 程序会查找文件名中包含货品ID的PDF文件。
        
        **Q: 处理很慢怎么办？**
        > A: 如果数据量大，请耐心等待。图片和条码文件较多时处理时间会更长。
        """)
    # 页脚
    st.divider()
    st.markdown(
        '<p style="text-align: center; color: #9ca3af;">出货单生成器 v2.0 | Powered by Streamlit</p>',
        unsafe_allow_html=True
    )
if __name__ == "__main__":
    main()